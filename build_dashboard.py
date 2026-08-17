#!/usr/bin/env python3
"""
Builds dashboard.html — an interactive, self-contained HTML dashboard from
financial.xlsx (Power BI Financial Sample) and ac-sample-data.xlsx (Chocolate Co. shipments).

Run:  python build_dashboard.py
Output: dashboard.html  (fully offline — Chart.js is inlined)
"""
import datetime
import json
import os
import openpyxl

OUT = "dashboard.html"
CHARTJS_FILE = "chart.umd.min.js"

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


# --------------------------------------------------------------------------
# 1. Load + aggregate the Power BI Financial Sample (financial.xlsx)
# --------------------------------------------------------------------------
def load_financial():
    wb = openpyxl.load_workbook("financial.xlsx", read_only=True, data_only=True)
    ws = wb["Sheet1"]
    recs = []
    for r in ws.iter_rows(values_only=True):
        if r[0] is None:
            continue
        try:
            units = float(r[4])
        except (TypeError, ValueError):
            continue
        d = r[12]
        recs.append({
            "segment": str(r[0]).strip(),
            "country": str(r[1]).strip(),
            "product": str(r[2]).strip(),
            "band": str(r[3]).strip(),
            "units": units,
            "sales": float(r[9]),
            "profit": float(r[11]),
            "year": int(r[15]),
            "month": d.month if isinstance(d, (datetime.datetime, datetime.date)) else 1,
        })
    wb.close()
    return recs


def agg_financial(recs):
    monthly = {}
    seg, cnt, prod, band = {}, {}, {}, {}
    for r in recs:
        key = (r["year"], r["month"])
        m = monthly.setdefault(key, {"sales": 0.0, "profit": 0.0, "units": 0.0})
        m["sales"] += r["sales"]
        m["profit"] += r["profit"]
        m["units"] += r["units"]

        s = seg.setdefault(r["segment"], {"sales": 0.0, "profit": 0.0, "units": 0.0})
        s["sales"] += r["sales"]; s["profit"] += r["profit"]; s["units"] += r["units"]

        c = cnt.setdefault(r["country"], {"sales": 0.0, "profit": 0.0, "units": 0.0})
        c["sales"] += r["sales"]; c["profit"] += r["profit"]; c["units"] += r["units"]

        p = prod.setdefault(r["product"], {"sales": 0.0, "profit": 0.0, "units": 0.0})
        p["sales"] += r["sales"]; p["profit"] += r["profit"]; p["units"] += r["units"]

        b = band.setdefault(r["band"], {"sales": 0.0, "profit": 0.0, "count": 0})
        b["sales"] += r["sales"]; b["profit"] += r["profit"]; b["count"] += 1

    def to_list(d, extra=None):
        out = []
        for name, v in d.items():
            item = {"name": name, "sales": round(v["sales"], 2),
                    "profit": round(v["profit"], 2), "units": round(v["units"], 1)}
            item["margin"] = round(item["profit"] / item["sales"], 4) if item["sales"] else 0
            out.append(item)
        out.sort(key=lambda x: x["sales"], reverse=True)
        return out

    monthly_list = []
    for (y, m), v in sorted(monthly.items()):
        monthly_list.append({
            "label": f"{MONTHS[m - 1]} {y % 100}",
            "sales": round(v["sales"], 2), "profit": round(v["profit"], 2),
            "units": round(v["units"], 1),
        })

    band_list = []
    for name, v in sorted(band.items(), key=lambda kv: kv[1]["sales"], reverse=True):
        band_list.append({
            "name": name, "sales": round(v["sales"], 2), "profit": round(v["profit"], 2),
            "count": v["count"],
            "margin": round(v["profit"] / v["sales"], 4) if v["sales"] else 0,
        })

    ts = sum(r["sales"] for r in recs)
    tp = sum(r["profit"] for r in recs)
    tu = sum(r["units"] for r in recs)
    # Row-level data (for the in-browser slicers + data explorer):
    # [year, month, segment, country, product, discount band, units, sales, profit]
    rows = [[r["year"], r["month"], r["segment"], r["country"], r["product"],
             r["band"], r["units"], r["sales"], r["profit"]] for r in recs]
    return {
        "kpis": {"sales": round(ts, 2), "profit": round(tp, 2), "units": round(tu, 1),
                 "margin": round(tp / ts, 4) if ts else 0, "rows": len(recs)},
        "monthly": monthly_list,
        "segments": to_list(seg),
        "countries": to_list(cnt),
        "products": to_list(prod),
        "bands": band_list,
        "rows": rows,
    }


# --------------------------------------------------------------------------
# 2. Load + aggregate Chocolate Co. shipments (ac-sample-data.xlsx)
# --------------------------------------------------------------------------
def load_shipments():
    wb = openpyxl.load_workbook("ac-sample-data.xlsx", read_only=True, data_only=True)

    # --- Dimension Data (lookups: product -> category/cost, geo -> region, person -> team)
    wsd = wb["Dimension Data"]
    drows = list(wsd.iter_rows(values_only=True))
    prod_cat, prod_cost, geo_region, person_team = {}, {}, {}, {}
    for r in drows[3:]:
        # Each lookup column is parsed independently: some trailing rows list
        # sales people without a product, so never skip a row wholesale.
        if r[1] is not None:
            p = str(r[1]).strip()
            prod_cat[p] = str(r[2]).strip() if r[2] else ""
            try:
                prod_cost[p] = float(r[3])
            except (TypeError, ValueError):
                prod_cost[p] = None
        g = str(r[7]).strip() if r[7] else ""
        if g:
            geo_region[g] = str(r[8]).strip() if r[8] else ""
        person = str(r[12]).strip() if r[12] else ""
        if person:
            person_team[person] = str(r[13]).strip() if r[13] else ""

    # --- Shipment Data (header at row 8: Sales Person, Geography, Product, Date, Sales, Boxes)
    ws = wb["Shipment Data"]
    rows = list(ws.iter_rows(values_only=True))
    recs = []
    for r in rows[8:]:
        if r[6] is None:
            continue
        geo = str(r[3]).strip()
        prod = str(r[4]).strip()
        person = str(r[2]).strip()
        d = r[5]
        recs.append({
            "person": person,
            "geo": geo,
            "region": geo_region.get(geo, ""),
            "product": prod,
            "cat": prod_cat.get(prod, ""),
            "team": person_team.get(person, "Unknown"),
            "date": d,
            "sales": float(r[6]),
            "boxes": float(r[7]) if r[7] else 0.0,
            "cost": prod_cost.get(prod),
        })
    wb.close()
    return recs


def agg_shipments(recs):
    monthly = {}
    geo, region, cat, team, prod, person = {}, {}, {}, {}, {}, {}
    est_profit_total = 0.0
    for r in recs:
        d = r["date"]
        y, m = (d.year, d.month) if isinstance(d, (datetime.datetime, datetime.date)) else (0, 0)
        key = (y, m)
        mm = monthly.setdefault(key, {"sales": 0.0, "boxes": 0.0})
        mm["sales"] += r["sales"]
        mm["boxes"] += r["boxes"]

        # Estimated gross profit = sales - (boxes x cost per box from the dimension table)
        est_profit = r["sales"] - r["boxes"] * (r["cost"] or 0.0)
        est_profit_total += est_profit

        g = geo.setdefault(r["geo"], {"sales": 0.0, "boxes": 0.0, "region": r["region"]})
        g["sales"] += r["sales"]; g["boxes"] += r["boxes"]

        if r["region"]:
            rg = region.setdefault(r["region"], {"sales": 0.0, "boxes": 0.0})
            rg["sales"] += r["sales"]; rg["boxes"] += r["boxes"]

        if r["cat"]:
            c = cat.setdefault(r["cat"], {"sales": 0.0, "boxes": 0.0, "profit": 0.0})
            c["sales"] += r["sales"]; c["boxes"] += r["boxes"]; c["profit"] += est_profit

        t = team.setdefault(r["team"], {"sales": 0.0, "boxes": 0.0})
        t["sales"] += r["sales"]; t["boxes"] += r["boxes"]

        p = prod.setdefault(r["product"], {"sales": 0.0, "boxes": 0.0, "cat": r["cat"], "profit": 0.0})
        p["sales"] += r["sales"]; p["boxes"] += r["boxes"]; p["profit"] += est_profit

        pe = person.setdefault(r["person"], {"sales": 0.0, "boxes": 0.0, "team": r["team"]})
        pe["sales"] += r["sales"]; pe["boxes"] += r["boxes"]

    def to_list(d, keys=("sales",)):
        out = []
        for name, v in d.items():
            item = {"name": name}
            for k in keys:
                item[k] = round(v[k], 2) if k != "boxes" else round(v[k], 1)
            out.append(item)
        out.sort(key=lambda x: x[keys[0]], reverse=True)
        return out

    monthly_list = []
    for (y, m), v in sorted(monthly.items()):
        monthly_list.append({"label": f"{MONTHS[m - 1]} {y % 100}",
                             "sales": round(v["sales"], 2), "boxes": round(v["boxes"], 1)})

    products = to_list(prod, ("sales", "boxes", "profit"))
    for p in products:
        p["cat"] = prod[p["name"]]["cat"]

    top_people = to_list(person, ("sales", "boxes"))
    for p in top_people:
        p["team"] = person[p["name"]]["team"]

    ts = sum(r["sales"] for r in recs)
    tb = sum(r["boxes"] for r in recs)
    # Row-level data (for the in-browser slicers + data explorer):
    # [ym, region, team, category, geo, product, person, sales, boxes, estProfit]
    rows = []
    for r in recs:
        d = r["date"]
        ym = d.year * 100 + d.month if isinstance(d, (datetime.datetime, datetime.date)) else 0
        est = r["sales"] - r["boxes"] * (r["cost"] or 0.0)
        rows.append([ym, r["region"], r["team"], r["cat"], r["geo"], r["product"],
                     r["person"], r["sales"], r["boxes"], round(est, 2)])
    return {
        "kpis": {"sales": round(ts, 2),
                 "boxes": round(tb, 1),
                 "shipments": len(recs),
                 "products": len(prod), "people": len(person), "geos": len(geo),
                 "estProfit": round(est_profit_total, 2),
                 "estMargin": round(est_profit_total / ts, 4) if ts else 0},
        "monthly": monthly_list,
        "regions": to_list(region, ("sales", "boxes")),
        "geos": to_list(geo, ("sales", "boxes"))[:8],
        "categories": to_list(cat, ("sales", "boxes", "profit")),
        "teams": to_list(team, ("sales", "boxes")),
        "products": products[:10],
        "topPeople": top_people[:10],
        "rows": rows,
    }


def build_insights(fin_all, ship):
    def top(lst):
        return lst[0] if lst else {"name": "-", "sales": 0, "margin": 0}

    seg = top(fin_all["segments"])
    cnt = top(fin_all["countries"])
    prd = top(fin_all["products"])
    team = top(ship["teams"])
    person = top(ship["topPeople"])
    region = top(ship["regions"])
    cat = top(ship["categories"])

    def best_worst(monthly):
        if not monthly:
            return None, None
        best = max(monthly, key=lambda x: x["sales"])
        worst = min(monthly, key=lambda x: x["sales"])
        return best, worst

    best_fin, worst_fin = best_worst(fin_all["monthly"])
    best_ship, _ = best_worst(ship["monthly"])

    y13 = sum(r[7] for r in fin_all["rows"] if r[0] == 2013)
    y14 = sum(r[7] for r in fin_all["rows"] if r[0] == 2014)
    yoy = round((y14 - y13) / y13, 4) if y13 else None

    k = fin_all["kpis"]
    sk = ship["kpis"]

    return {
        "topSegment": {"name": seg["name"], "value": seg["sales"],
                       "share": round(seg["sales"] / fin_all["kpis"]["sales"], 3)},
        "topCountry": {"name": cnt["name"], "value": cnt["sales"],
                       "share": round(cnt["sales"] / fin_all["kpis"]["sales"], 3)},
        "topProduct": {"name": prd["name"], "value": prd["sales"],
                       "share": round(prd["sales"] / fin_all["kpis"]["sales"], 3)},
        "topTeam": {"name": team["name"], "value": team["sales"]},
        "topPerson": {"name": person["name"], "value": person["sales"]},
        "topRegion": {"name": region["name"], "value": region["sales"]},
        "topCategory": {"name": cat["name"], "value": cat["sales"]},
        # ---- additional derived info ----
        "bestMonth": {"label": best_fin["label"], "value": best_fin["sales"]},
        "worstMonth": {"label": worst_fin["label"], "value": worst_fin["sales"]},
        "bestShipMonth": {"label": best_ship["label"], "value": best_ship["sales"]},
        "yoy": yoy,
        "y13": round(y13, 2),
        "y14": round(y14, 2),
        "avgOrder": round(k["sales"] / k["rows"], 2) if k["rows"] else 0,
        "avgSalePerUnit": round(k["sales"] / k["units"], 2) if k["units"] else 0,
        "avgShipment": round(sk["sales"] / sk["shipments"], 2) if sk["shipments"] else 0,
        "avgBoxValue": round(sk["sales"] / sk["boxes"], 2) if sk["boxes"] else 0,
        "topPersonShare": round(person["sales"] / sk["sales"], 3) if sk["sales"] else 0,
        "topRegionShare": round(region["sales"] / sk["sales"], 3) if sk["sales"] else 0,
        "topCatShare": round(cat["sales"] / sk["sales"], 3) if sk["sales"] else 0,
        "topProdMargin": prd["margin"],
        "bestMarginSeg": max(fin_all["segments"], key=lambda x: x["margin"]).get("name", "-"),
    }


def build_combined(fin_all, ship):
    """Cross-dataset view: combined top products, indexed monthly trend, totals."""
    products = []
    for p in fin_all["products"]:
        products.append({"name": p["name"], "dataset": "Power BI",
                         "sales": p["sales"], "profit": p["profit"],
                         "volume": p["units"]})
    for p in ship["products"]:
        products.append({"name": p["name"], "dataset": "Chocolate",
                         "sales": p["sales"], "profit": p["profit"],
                         "volume": p["boxes"]})
    products.sort(key=lambda x: x["sales"], reverse=True)
    products = products[:12]

    fin_sales = fin_all["kpis"]["sales"]
    choc_sales = ship["kpis"]["sales"]
    choc_profit = ship["kpis"]["estProfit"]
    total_sales = fin_sales + choc_sales
    total_profit = fin_all["kpis"]["profit"] + choc_profit
    return {
        "products": products,
        "totals": {
            "sales": round(total_sales, 2),
            "profit": round(total_profit, 2),
            "margin": round(total_profit / total_sales, 4) if total_sales else 0,
            "volume": round(fin_all["kpis"]["units"] + ship["kpis"]["boxes"], 1),
            "finProfit": round(fin_all["kpis"]["profit"], 2),
            "finMargin": fin_all["kpis"]["margin"],
            "chocProfit": round(choc_profit, 2),
            "chocMargin": ship["kpis"]["estMargin"],
        },
    }


# --------------------------------------------------------------------------
# 3. HTML template (placeholders: __CHARTJS__, __DATA__)
# --------------------------------------------------------------------------
TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta property="og:title" content="Interactive Data Analytics Dashboard — Excel to Web">
<meta property="og:description" content="Power BI-style analytics dashboard built from raw Excel data: 4 tabs plus a data explorer, KPI cards with sparklines, donuts with centre totals, country radar, goal rings, live slicers and year filters. Built with Python, Chart.js and GitHub Pages.">
<meta property="og:image" content="https://calvinokoth9528-cloud.github.io/excel-powerbi-dashboard/preview.png">
<meta property="og:url" content="https://calvinokoth9528-cloud.github.io/excel-powerbi-dashboard/dashboard.html">
<title>Data Analytics Dashboard — Attachment Portfolio</title>
<style>
  :root {
    --bg: #f3f5fa;
    --card: #ffffff;
    --line: #e8ebf3;
    --text: #1d2537;
    --muted: #7c869c;
    --faint: #a8b0c2;
    --blue: #4c6fff;
    --blue-soft: #e8edff;
    --green: #23b26d;
    --green-soft: #e4f7ee;
    --orange: #f2994a;
    --orange-soft: #fdf0e3;
    --red: #ef5b5b;
    --red-soft: #fdeaea;
    --pink: #f472b6;
    --pink-soft: #fdebf4;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    background: var(--bg);
    color: var(--text);
    font-family: "Segoe UI", system-ui, -apple-system, "Helvetica Neue", Arial, sans-serif;
    min-height: 100vh;
  }
  header {
    position: sticky; top: 0; z-index: 20;
    background: rgba(255, 255, 255, 0.88);
    backdrop-filter: blur(10px);
    border-bottom: 1px solid var(--line);
  }
  .bar {
    max-width: 1280px; margin: 0 auto; padding: 14px 28px;
    display: flex; align-items: center; justify-content: space-between; gap: 16px; flex-wrap: wrap;
  }
  .brand { display: flex; align-items: center; gap: 12px; }
  .logo {
    width: 40px; height: 40px; border-radius: 12px; flex: none;
    background: linear-gradient(135deg, var(--blue), #8b5cf6);
    color: #fff; display: flex; align-items: center; justify-content: center;
    font-weight: 800; font-size: 17px;
    box-shadow: 0 6px 16px rgba(76, 111, 255, 0.35);
  }
  .brand h1 { font-size: 18px; font-weight: 700; letter-spacing: 0.2px; }
  .brand p { font-size: 12.5px; color: var(--muted); margin-top: 2px; }
  nav { display: flex; flex-wrap: wrap; gap: 4px; background: var(--bg); padding: 5px; border-radius: 12px; border: 1px solid var(--line); }
  .tab-btn {
    background: none; border: none; color: var(--muted); cursor: pointer;
    font-size: 13.5px; font-weight: 600; padding: 9px 16px; border-radius: 9px; transition: all .15s;
    font-family: inherit;
  }
  .tab-btn:hover { color: var(--text); background: #fff; }
  .tab-btn.active { background: #fff; color: var(--blue); box-shadow: 0 2px 10px rgba(28, 37, 55, 0.10); }
  main { max-width: 1280px; margin: 0 auto; padding: 24px 28px 40px; }
  .tab { display: none; }
  .tab.active { display: block; }
  .kpi-row { display: grid; grid-template-columns: repeat(auto-fit, minmax(215px, 1fr)); gap: 16px; margin-bottom: 20px; }
  .kpi {
    background: var(--card); border: 1px solid var(--line); border-radius: 16px; padding: 15px 18px 8px;
    box-shadow: 0 6px 20px rgba(28, 37, 55, 0.05); position: relative; display: flex; flex-direction: column;
  }
  .kpi-top { display: flex; align-items: center; justify-content: space-between; }
  .kpi .label { font-size: 12px; color: var(--muted); font-weight: 600; }
  .kpi .icon {
    width: 34px; height: 34px; border-radius: 10px; flex: none;
    display: flex; align-items: center; justify-content: center; font-size: 16px;
  }
  .icon.blue { background: var(--blue-soft); }
  .icon.green { background: var(--green-soft); }
  .icon.orange { background: var(--orange-soft); }
  .icon.pink { background: var(--pink-soft); }
  .kpi .value { font-size: 26px; font-weight: 700; margin-top: 8px; letter-spacing: -0.6px; }
  .kpi .delta { font-size: 11.5px; font-weight: 600; margin-top: 5px; display: flex; align-items: center; gap: 4px; }
  .delta.up { color: var(--green); }
  .delta.down { color: var(--red); }
  .delta.flat { color: var(--muted); font-weight: 500; }
  .kpi .spark { width: 100%; height: 34px; margin-top: 8px; display: block; }
  .ring-card {
    background: var(--card); border: 1px solid var(--line); border-radius: 16px; padding: 15px 18px;
    box-shadow: 0 6px 20px rgba(28, 37, 55, 0.05); display: flex; flex-direction: column;
  }
  .ring-wrap { position: relative; width: 96px; height: 96px; margin: 8px auto 2px; }
  .ring {
    width: 100%; height: 100%; border-radius: 50%;
    background: conic-gradient(var(--blue) calc(var(--p) * 1%), #edf0f7 0);
    display: flex; align-items: center; justify-content: center;
  }
  .ring::before { content: ""; position: absolute; inset: 11px; background: #fff; border-radius: 50%; }
  .ring b { position: relative; z-index: 1; font-size: 20px; }
  .ring-cap { text-align: center; font-size: 11px; color: var(--muted); font-weight: 600; text-transform: uppercase; letter-spacing: 0.6px; margin-top: 4px; }
  .grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 16px; }
  .span2 { grid-column: span 2; }
  .card {
    background: var(--card); border: 1px solid var(--line); border-radius: 16px; padding: 18px 20px;
    box-shadow: 0 6px 20px rgba(28, 37, 55, 0.05);
  }
  .card h3 { font-size: 13.5px; font-weight: 700; color: var(--text); margin-bottom: 14px; letter-spacing: 0.2px; display: flex; align-items: center; gap: 8px; }
  .card h3::before { content: ""; width: 8px; height: 8px; border-radius: 3px; background: var(--blue); flex: none; }
  .chart-wrap { position: relative; height: 300px; }
  .chart-wrap.tall { height: 340px; }
  .seg { display: flex; gap: 8px; margin: 0 0 16px; flex-wrap: wrap; }
  .seg-btn {
    background: #fff; border: 1px solid var(--line); color: var(--muted);
    font-size: 12.5px; font-weight: 600; padding: 7px 14px; border-radius: 999px; cursor: pointer; transition: all .15s; font-family: inherit;
  }
  .seg-btn:hover { color: var(--text); border-color: #cdd3e2; }
  .seg-btn.active { background: var(--blue); color: #fff; border-color: var(--blue); box-shadow: 0 4px 12px rgba(76, 111, 255, 0.30); }
  /* ---- Power BI-style slicers ---- */
  .pill {
    display: inline-block; font-size: 11.5px; font-weight: 700; letter-spacing: .4px;
    color: #fff; background: linear-gradient(90deg, var(--blue), #8b5cf6);
    padding: 6px 13px; border-radius: 999px; text-transform: uppercase;
    box-shadow: 0 4px 12px rgba(76, 111, 255, 0.3);
  }
  .filters {
    background: var(--card); border: 1px solid var(--line); border-radius: 16px;
    padding: 16px 20px; margin-bottom: 20px; box-shadow: 0 6px 20px rgba(28, 37, 55, 0.05);
  }
  .filters h3 { font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: 1.2px; color: var(--muted); margin-bottom: 12px; }
  .frow { display: flex; gap: 22px; flex-wrap: wrap; align-items: flex-end; }
  .fgroup { display: flex; flex-direction: column; gap: 7px; }
  .fgroup label { font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: .8px; color: var(--faint); }
  .chips { display: flex; gap: 7px; flex-wrap: wrap; }
  .chip {
    background: #fff; border: 1px solid var(--line); color: var(--muted); cursor: pointer;
    font-size: 12.5px; font-weight: 600; padding: 7px 13px; border-radius: 999px;
    transition: all .15s; font-family: inherit; user-select: none;
  }
  .chip .n { font-weight: 700; opacity: .75; }
  .chip:hover { color: var(--text); border-color: #cdd3e2; }
  .chip.on { background: var(--blue); color: #fff; border-color: var(--blue); box-shadow: 0 4px 12px rgba(76, 111, 255, 0.28); }
  .chip.on .n { opacity: .9; }
  .yr { display: flex; align-items: center; gap: 8px; }
  .yr input {
    width: 74px; padding: 8px 10px; border: 1px solid var(--line); border-radius: 9px;
    font-size: 13px; font-family: inherit; font-weight: 600; color: var(--text); background: #fff;
  }
  .yr span { color: var(--muted); font-size: 12px; }
  .reset {
    background: none; border: 1px solid var(--line); color: var(--muted); cursor: pointer;
    font-size: 12.5px; font-weight: 600; padding: 8px 16px; border-radius: 999px;
    transition: all .15s; font-family: inherit; margin-left: auto;
  }
  .reset:hover { color: var(--red); border-color: var(--red); }
  .filter-note { font-size: 11.5px; color: var(--faint); margin-top: 10px; }
  /* ---- live insights ---- */
  .insight-list { display: flex; flex-direction: column; gap: 10px; }
  .insight-row {
    display: flex; gap: 11px; align-items: flex-start; background: #fafbfe;
    border: 1px solid var(--line); border-radius: 12px; padding: 12px 14px;
  }
  .insight-row .em { font-size: 17px; flex: none; }
  .insight-row .it b { font-size: 13px; color: var(--text); }
  .insight-row .it span { font-size: 12.5px; color: var(--muted); line-height: 1.5; display: block; margin-top: 2px; }
  /* ---- data explorer ---- */
  .exp-search {
    width: 100%; padding: 10px 14px; border: 1px solid var(--line); border-radius: 10px;
    font-size: 13.5px; font-family: inherit; background: #fff; color: var(--text); margin-bottom: 12px;
  }
  .exp-table-wrap { max-height: 430px; overflow: auto; border: 1px solid var(--line); border-radius: 12px; }
  .exp-table-wrap table { font-size: 12.5px; }
  .exp-table-wrap thead th { position: sticky; top: 0; background: #fafbfe; z-index: 1; }
  .tag {
    display: inline-block; font-size: 11px; font-weight: 700; letter-spacing: .5px; color: var(--blue);
    background: var(--blue-soft); padding: 4px 10px; border-radius: 999px;
  }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  th { text-align: left; font-size: 11px; text-transform: uppercase; letter-spacing: 0.8px; color: var(--faint); font-weight: 600; padding: 8px 10px; border-bottom: 1px solid var(--line); }
  td { padding: 9px 10px; border-bottom: 1px solid rgba(28, 37, 55, 0.05); color: var(--text); }
  tr:last-child td { border-bottom: none; }
  td.num, th.num { text-align: right; font-variant-numeric: tabular-nums; }
  .rank { display: inline-flex; align-items: center; justify-content: center; width: 22px; height: 22px; border-radius: 7px; background: var(--bg); font-size: 11.5px; font-weight: 700; color: var(--muted); }
  tr.first .rank { background: var(--blue); color: #fff; }
  tr.first td:first-child { color: var(--blue); }
  .insights { grid-template-columns: repeat(auto-fit, minmax(240px, 1fr)); }
  .insight { padding: 16px 18px; }
  .insight .i-label { font-size: 11.5px; text-transform: uppercase; letter-spacing: 1px; color: var(--faint); font-weight: 600; }
  .insight .i-value { font-size: 21px; font-weight: 700; margin-top: 7px; color: var(--blue); }
  .insight .i-sub { font-size: 12px; color: var(--muted); margin-top: 4px; }
  .goals { display: flex; flex-direction: column; gap: 13px; padding-top: 6px; }
  .goal .g-top { display: flex; justify-content: space-between; font-size: 12.5px; margin-bottom: 5px; }
  .goal .g-top b { font-weight: 600; }
  .goal .g-top span { color: var(--muted); font-variant-numeric: tabular-nums; }
  .goal .g-bar { height: 8px; border-radius: 999px; background: #edf0f7; overflow: hidden; }
  .goal .g-fill { height: 100%; border-radius: 999px; background: linear-gradient(90deg, var(--blue), #8b5cf6); }
  .goal:nth-child(2n) .g-fill { background: linear-gradient(90deg, var(--green), #4ade80); }
  .goal:nth-child(3n) .g-fill { background: linear-gradient(90deg, var(--orange), #f6c26b); }
  footer { max-width: 1280px; margin: 0 auto; padding: 0 28px 34px; color: var(--faint); font-size: 12px; line-height: 1.6; }
  footer b { color: var(--muted); font-weight: 600; }
  @media (max-width: 900px) {
    .grid { grid-template-columns: 1fr; }
    .span2 { grid-column: span 1; }
    .bar { padding: 14px 18px; }
    main { padding: 18px 18px 28px; }
  }
</style>
</head>
<body>
<header>
  <div class="bar">
    <div class="brand">
      <div class="logo">CO</div>
      <div>
        <h1>Data Analytics Dashboard</h1>
        <p>Attachment portfolio · Calvin Okoth · Excel → Power BI workflow</p>
      </div>
    </div>
    <div style="display:flex;align-items:center;gap:14px;flex-wrap:wrap">
      <span class="pill">Interactive — slicers filter every chart</span>
      <nav>
        <button class="tab-btn active" data-tab="overview">Overview</button>
        <button class="tab-btn" data-tab="financial">Power BI · Financial</button>
        <button class="tab-btn" data-tab="shipments">Chocolate Co. · Shipments</button>
        <button class="tab-btn" data-tab="combined">Combined Report</button>
        <button class="tab-btn" data-tab="data">Data Explorer</button>
      </nav>
    </div>
  </div>
</header>

<main>
  <!-- ================= OVERVIEW ================= -->
  <section class="tab active" id="tab-overview">
    <div class="kpi-row" id="ov-kpis"></div>
    <div class="grid">
      <div class="card span2"><h3>Financial Sample — Monthly Sales &amp; Profit (2013–2014)</h3><div class="chart-wrap" id="wrap-ov-fin"></div></div>
      <div class="card"><h3>Financial Sales Share by Segment</h3><div class="chart-wrap" id="wrap-ov-seg"></div></div>
      <div class="card"><h3>Financial Sales by Country</h3><div class="chart-wrap" id="wrap-ov-cnt"></div></div>
      <div class="card span2"><h3>Chocolate Co. — Monthly Shipment Sales (Feb 2023 – Feb 2024)</h3><div class="chart-wrap" id="wrap-ov-ship"></div></div>
      <div class="card"><h3>Shipment Sales by Region</h3><div class="chart-wrap" id="wrap-ov-reg"></div></div>
      <div class="card"><h3>Key Findings</h3><div id="ov-findings"></div></div>
    </div>
    <div class="insights grid" id="ov-growth"></div>
  </section>

  <!-- ================= FINANCIAL ================= -->
  <section class="tab" id="tab-financial">
    <div class="filters">
      <h3>Filters (Power BI-style slicers)</h3>
      <div class="frow">
        <div class="fgroup">
          <label>Segment</label>
          <div class="chips" id="chip-seg"></div>
        </div>
        <div class="fgroup">
          <label>Country</label>
          <div class="chips" id="chip-cnt"></div>
        </div>
        <div class="fgroup">
          <label>Year</label>
          <div class="yr">
            <input type="number" id="finYrMin" min="2013" max="2014" value="2013">
            <span>to</span>
            <input type="number" id="finYrMax" min="2013" max="2014" value="2014">
          </div>
        </div>
        <button class="reset" id="btnResetFin">Reset all</button>
      </div>
      <div class="filter-note" id="finFilterNote"></div>
    </div>
    <div class="kpi-row" id="fin-kpis"></div>
    <div class="grid">
      <div class="card span2"><h3>Monthly Sales &amp; Profit</h3><div class="chart-wrap" id="wrap-fin-month"></div></div>
      <div class="card"><h3>Sales Share by Segment</h3><div class="chart-wrap" id="wrap-fin-seg"></div></div>
      <div class="card"><h3>Country Performance Profile</h3><div class="chart-wrap" id="wrap-fin-radar"></div></div>
      <div class="card"><h3>Top 5 Products by Sales</h3><div id="fin-goals" class="goals"></div></div>
      <div class="card"><h3>Sales by Country</h3><div class="chart-wrap" id="wrap-fin-cnt"></div></div>
      <div class="card"><h3>Sales by Product</h3><div class="chart-wrap" id="wrap-fin-prod"></div></div>
      <div class="card"><h3>Discount Bands — Profit Margin &amp; Share</h3><div class="chart-wrap" id="wrap-fin-band"></div></div>
      <div class="card span2"><h3>Live Insights <span class="tag" style="background:var(--green-soft);color:var(--green)">update as you filter</span></h3><div id="fin-insights"></div></div>
    </div>
  </section>

  <!-- ================= SHIPMENTS ================= -->
  <section class="tab" id="tab-shipments">
    <div class="filters">
      <h3>Filters (Power BI-style slicers)</h3>
      <div class="frow">
        <div class="fgroup">
          <label>Region</label>
          <div class="chips" id="chip-reg"></div>
        </div>
        <div class="fgroup">
          <label>Team</label>
          <div class="chips" id="chip-team"></div>
        </div>
        <div class="fgroup">
          <label>Category</label>
          <div class="chips" id="chip-cat"></div>
        </div>
        <button class="reset" id="btnResetShip">Reset all</button>
      </div>
      <div class="filter-note" id="shipFilterNote"></div>
    </div>
    <div class="kpi-row" id="ship-kpis"></div>
    <div class="grid">
      <div class="card span2"><h3>Monthly Sales &amp; Boxes Shipped</h3><div class="chart-wrap" id="wrap-ship-month"></div></div>
      <div class="card"><h3>Sales by Region</h3><div class="chart-wrap" id="wrap-ship-reg"></div></div>
      <div class="card"><h3>Sales by Geography</h3><div class="chart-wrap" id="wrap-ship-geo"></div></div>
      <div class="card"><h3>Sales by Product Category</h3><div class="chart-wrap" id="wrap-ship-cat"></div></div>
      <div class="card"><h3>Sales by Team</h3><div class="chart-wrap" id="wrap-ship-team"></div></div>
      <div class="card span2"><h3>Top 10 Sales People</h3><div id="ship-top-people"></div></div>
      <div class="card span2"><h3>Live Insights <span class="tag" style="background:var(--green-soft);color:var(--green)">update as you filter</span></h3><div id="ship-insights"></div></div>
    </div>
  </section>

  <!-- ================= COMBINED REPORT ================= -->
  <section class="tab" id="tab-combined">
    <div class="seg" id="comb-metric-seg">
      <button class="seg-btn active" data-metric="sales">Sales</button>
      <button class="seg-btn" data-metric="profit">Profit</button>
      <button class="seg-btn" data-metric="volume">Volume</button>
    </div>
    <div class="kpi-row" id="comb-kpis"></div>
    <div class="grid">
      <div class="card"><h3>Monthly Sales — Power BI Sample (2013–14)</h3><div class="chart-wrap" id="wrap-comb-trend-fin"></div></div>
      <div class="card"><h3>Monthly Sales — Chocolate Co. (2023–24)</h3><div class="chart-wrap" id="wrap-comb-trend-choc"></div></div>
      <div class="card"><h3>Revenue by Segment — Power BI Sample</h3><div class="chart-wrap" id="wrap-comb-seg"></div></div>
      <div class="card"><h3>Revenue by Region — Chocolate Co.</h3><div class="chart-wrap" id="wrap-comb-reg"></div></div>
      <div class="card span2"><h3>Top 12 Products Across Both Datasets</h3><div class="chart-wrap tall" id="wrap-comb-prod"></div></div>
      <div class="card"><h3>Profit Margin by Segment — Power BI Sample</h3><div class="chart-wrap" id="wrap-comb-margin"></div></div>
      <div class="card"><h3>Est. Profit by Category — Chocolate Co.</h3><div class="chart-wrap" id="wrap-comb-cat"></div></div>
      <div class="card span2"><h3>Both Datasets at a Glance</h3><div id="comb-table"></div></div>
    </div>
  </section>

  <!-- ================= DATA EXPLORER ================= -->
  <section class="tab" id="tab-data">
    <div class="kpi-row" id="data-kpis"></div>
    <div class="grid">
      <div class="card span2">
        <h3>Financial Sample — full dataset <span class="tag" id="dtFinCount">—</span></h3>
        <div class="cap" style="font-size:12px;color:var(--muted);margin-bottom:10px">Search any field · rows follow the Financial tab slicers</div>
        <input type="text" class="exp-search" id="dtFinSearch" placeholder="Search segment, country, product…">
        <div class="exp-table-wrap"><table><thead><tr><th>Year</th><th>Month</th><th>Segment</th><th>Country</th><th>Product</th><th>Band</th><th class="num">Units</th><th class="num">Sales</th><th class="num">Profit</th></tr></thead><tbody id="dtFinBody"></tbody></table></div>
      </div>
      <div class="card span2">
        <h3>Chocolate Co. — full dataset <span class="tag" id="dtShipCount">—</span></h3>
        <div class="cap" style="font-size:12px;color:var(--muted);margin-bottom:10px">Search any field · rows follow the Shipments tab slicers</div>
        <input type="text" class="exp-search" id="dtShipSearch" placeholder="Search region, product, sales person…">
        <div class="exp-table-wrap"><table><thead><tr><th>Period</th><th>Region</th><th>Team</th><th>Category</th><th>Geography</th><th>Product</th><th>Person</th><th class="num">Sales</th><th class="num">Boxes</th><th class="num">Est. profit</th></tr></thead><tbody id="dtShipBody"></tbody></table></div>
      </div>
    </div>
  </section>
</main>

<footer>
  <b>Sources:</b> <b>financial.xlsx</b> — Power BI Financial Sample (700 rows · 2013–2014) &nbsp;·&nbsp; <b>ac-sample-data.xlsx</b> — Chocolate Co. Shipment Data (6,113 rows · Feb 2023 – Feb 2024) plus Dimension &amp; Calendar tables.
  Built with Python + Chart.js (v4, embedded — works fully offline). Hover any chart for details.
</footer>

<script>
/*__CHARTJS__*/
</script>
<script>
"use strict";
const DATA = /*__DATA__*/;

const TEAL = "#2dd4bf", GOLD = "#f5b942", BLUE = "#60a5fa", PURPLE = "#a78bfa",
      ROSE = "#fb7185", GREEN = "#4ade80", ORANGE = "#fb923c", CYAN = "#22d3ee",
      PINK = "#f472b6", LIME = "#a3e635";
const PALETTE = [TEAL, GOLD, BLUE, PURPLE, ROSE, GREEN, ORANGE, CYAN, PINK, LIME];
const F = new Intl.NumberFormat("en-US");

function money(n) {
  if (n >= 1e9) return "$" + (n / 1e9).toFixed(2) + "B";
  if (n >= 1e6) return "$" + (n / 1e6).toFixed(1) + "M";
  if (n >= 1e3) return "$" + (n / 1e3).toFixed(0) + "K";
  return "$" + Math.round(n).toLocaleString();
}
function num(n) {
  if (n >= 1e6) return (n / 1e6).toFixed(1) + "M";
  if (n >= 1e3) return (n / 1e3).toFixed(0) + "K";
  return Math.round(n).toLocaleString();
}
function pct(x) { return (x * 100).toFixed(1) + "%"; }

Chart.defaults.color = "#7c869c";
Chart.defaults.borderColor = "rgba(28, 37, 55, 0.08)";
Chart.defaults.font.family = "'Segoe UI', system-ui, -apple-system, sans-serif";
Chart.defaults.plugins.legend.labels.usePointStyle = true;
Chart.defaults.plugins.legend.labels.boxWidth = 8;
Chart.defaults.plugins.legend.labels.padding = 14;
Chart.defaults.plugins.tooltip.backgroundColor = "#ffffff";
Chart.defaults.plugins.tooltip.borderColor = "rgba(28, 37, 55, 0.10)";
Chart.defaults.plugins.tooltip.borderWidth = 1;
Chart.defaults.plugins.tooltip.titleColor = "#1d2537";
Chart.defaults.plugins.tooltip.bodyColor = "#4a5568";
Chart.defaults.plugins.tooltip.padding = 10;
Chart.defaults.plugins.tooltip.cornerRadius = 10;
Chart.defaults.animation.duration = 550;

/* Draw a big number + label in the middle of a doughnut (Excel dashboard style) */
const centerText = {
  id: "centerText",
  afterDraw(chart) {
    const cfg = chart.options.plugins && chart.options.plugins.centerText;
    if (!cfg) return;
    const { ctx, chartArea } = chart;
    const cx = (chartArea.left + chartArea.right) / 2;
    const cy = (chartArea.top + chartArea.bottom) / 2;
    ctx.save();
    ctx.textAlign = "center";
    ctx.textBaseline = "middle";
    ctx.fillStyle = "#1d2537";
    ctx.font = "700 22px 'Segoe UI', sans-serif";
    ctx.fillText(cfg.value, cx, cy - 8);
    ctx.fillStyle = "#7c869c";
    ctx.font = "600 9.5px 'Segoe UI', sans-serif";
    ctx.fillText(String(cfg.label).toUpperCase(), cx, cy + 11);
    ctx.restore();
  },
};
Chart.register(centerText);

const charts = {};
function makeChart(id, cfg) {
  if (charts[id]) charts[id].destroy();
  const el = canvas(id);
  if (!el) return;
  charts[id] = new Chart(el, cfg);
}

function tooltipFor(fmt) {
  return {
    callbacks: {
      label(ctx) {
        const v = ctx.parsed && typeof ctx.parsed === "object"
          ? (ctx.parsed.y ?? ctx.parsed.x ?? ctx.parsed.r) : ctx.parsed;
        const lbl = ctx.dataset.label ? ctx.dataset.label + ": " : "";
        return lbl + (fmt === "pct" ? pct(v) : fmt === "int" ? num(v) : money(v));
      }
    }
  };
}
function moneyTicks(fmt) {
  return { ticks: { callback: v => (fmt === "pct" ? pct(v) : fmt === "int" ? num(v) : money(v)) } };
}

function lineCfg(id, labels, datasets, extra) {
  const ds = datasets.map((d, i) => ({
    label: d.label, data: d.data,
    borderColor: d.color, backgroundColor: d.color,
    yAxisID: d.axis || "y",
    tension: 0.35, fill: false, borderWidth: 2.5, pointRadius: 2, pointHoverRadius: 5,
  }));
  const scales = {
    x: { grid: { display: false } },
    y: { position: "left", ...moneyTicks(extra.fmt || "money") },
  };
  if (extra.second) {
    scales.y1 = { position: "right", grid: { drawOnChartArea: false }, ...moneyTicks(extra.fmt2 || "money") };
  }
  return {
    type: "line",
    data: { labels, datasets: ds },
    options: {
      responsive: true, maintainAspectRatio: false,
      interaction: { mode: "index", intersect: false },
      plugins: { legend: { position: "top", align: "end" }, tooltip: tooltipFor(extra.fmt || "money") },
      scales,
    },
  };
}

function doughnutCfg(labels, values, fmt, center) {
  return {
    type: "doughnut",
    data: {
      labels,
      datasets: [{ data: values, backgroundColor: labels.map((_, i) => PALETTE[i % PALETTE.length]),
                   borderColor: "#ffffff", borderWidth: 3 }],
    },
    options: {
      responsive: true, maintainAspectRatio: false, cutout: "68%",
      plugins: {
        legend: { position: "bottom" },
        tooltip: tooltipFor(fmt || "money"),
        centerText: center || false,
      },
    },
  };
}

function radarCfg(labels, datasets) {
  return {
    type: "radar",
    data: {
      labels,
      datasets: datasets.map((ds, i) => ({
        label: ds.label, data: ds.data,
        backgroundColor: PALETTE[i] + "2e",
        borderColor: PALETTE[i], borderWidth: 2,
        pointBackgroundColor: PALETTE[i], pointRadius: 3, pointHoverRadius: 5,
      })),
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      scales: {
        r: {
          min: 0, max: 100, ticks: { display: false },
          grid: { color: "rgba(28, 37, 55, 0.08)" },
          angleLines: { color: "rgba(28, 37, 55, 0.08)" },
          pointLabels: { color: "#4a5568", font: { size: 11, weight: 600 } },
        },
      },
      plugins: {
        legend: { position: "top", align: "end" },
        tooltip: { callbacks: { label: ctx => ctx.dataset.label + ": " + ctx.parsed.r.toFixed(1) + "% share" } },
      },
    },
  };
}

function shareNorm(list, key) {
  const tot = list.reduce((s, x) => s + x[key], 0) || 1;
  return list.map(x => +(x[key] / tot * 100).toFixed(1));
}

function barCfg(labels, values, opts) {
  const horizontal = opts && opts.horizontal;
  const fmt = (opts && opts.fmt) || "money";
  const colors = opts && opts.colors ? opts.colors : labels.map((_, i) => PALETTE[i % PALETTE.length]);
  const data = { labels, datasets: [{ data: values, backgroundColor: colors, borderRadius: 5, borderSkipped: false, maxBarThickness: 34 }] };
  const scales = horizontal
    ? { x: { grid: { display: false }, ...moneyTicks(fmt) }, y: { grid: { display: false } } }
    : { x: { grid: { display: false } }, y: { grid: { color: "rgba(28,37,55,0.06)" }, ...moneyTicks(fmt) } };
  return {
    type: "bar",
    data,
    options: {
      indexAxis: horizontal ? "y" : "x",
      responsive: true, maintainAspectRatio: false,
      plugins: { legend: { display: false }, tooltip: tooltipFor(fmt) },
      scales,
    },
  };
}

function canvas(id) {
  const wrap = document.getElementById("wrap-" + id);
  if (!wrap) return null;
  wrap.innerHTML = '<canvas id="' + id + '"></canvas>';
  return wrap.querySelector("canvas");
}

/* ---------- KPI card builder (Excel dashboard style: icon + delta + sparkline) ---------- */
const __sparks = [];
let __sparkSeq = 0;

function kpi(label, value, delta, opt) {
  opt = opt || {};
  const d = delta
    ? '<div class="delta ' + (delta.cls || "flat") + '">' + (delta.arrow || "") + delta.text + "</div>"
    : "";
  let spark = "";
  if (opt.spark && opt.spark.length > 1) {
    const id = "spark-" + (++__sparkSeq);
    __sparks.push({ id, data: opt.spark, color: opt.color || "#4c6fff" });
    spark = '<canvas class="spark" id="' + id + '"></canvas>';
  }
  return '<div class="kpi"><div class="kpi-top"><div class="label">' + label + "</div>" +
    (opt.icon ? '<div class="icon ' + (opt.iconCls || "blue") + '">' + opt.icon + "</div>" : "") + "</div>" +
    '<div class="value">' + value + "</div>" + d + spark + "</div>";
}

function sparkline(canvas, data, color) {
  const dpr = window.devicePixelRatio || 1;
  const w = canvas.clientWidth, h = canvas.clientHeight;
  if (!w || !h) return;
  canvas.width = w * dpr; canvas.height = h * dpr;
  const ctx = canvas.getContext("2d");
  ctx.scale(dpr, dpr);
  if (!data || data.length < 2) return;
  const min = Math.min(...data), max = Math.max(...data);
  const pad = 2, span = (max - min) || 1;
  const pts = data.map((v, i) => [
    pad + (i / (data.length - 1)) * (w - pad * 2),
    h - pad - ((v - min) / span) * (h - pad * 2),
  ]);
  ctx.beginPath();
  ctx.moveTo(pts[0][0], pts[0][1]);
  for (let i = 1; i < pts.length; i++) ctx.lineTo(pts[i][0], pts[i][1]);
  ctx.strokeStyle = color; ctx.lineWidth = 2; ctx.lineJoin = "round"; ctx.lineCap = "round";
  ctx.stroke();
  ctx.lineTo(pts[pts.length - 1][0], h - 1); ctx.lineTo(pts[0][0], h - 1); ctx.closePath();
  const g = ctx.createLinearGradient(0, 0, 0, h);
  g.addColorStop(0, color + "33"); g.addColorStop(1, color + "00");
  ctx.fillStyle = g; ctx.fill();
}

function flushSparks() {
  __sparks.forEach(s => { const el = document.getElementById(s.id); if (el) sparkline(el, s.data, s.color); });
  __sparks.length = 0;
}

function lastDelta(monthly, key) {
  const a = monthly[monthly.length - 2], b = monthly[monthly.length - 1];
  if (!a || !b || !a[key]) return null;
  return (b[key] - a[key]) / a[key];
}

function deltaHtml(d) {
  if (d === null || d === undefined) return null;
  const up = d >= 0;
  return { text: (up ? "▲ " : "▼ ") + Math.abs(d * 100).toFixed(1) + "% vs prev month", cls: up ? "up" : "down" };
}

function ringCard(label, percent, value) {
  return '<div class="ring-card"><div class="label" style="font-size:12px;color:var(--muted);font-weight:600">' + label + '</div>' +
    '<div class="ring-wrap"><div class="ring" style="--p:' + percent + '"><b>' + value + "</b></div></div>" +
    '<div class="ring-cap">' + label + " goal</div></div>";
}

function goalsHtml(items, fmt) {
  const max = items[0] ? items[0].value : 1;
  return '<div class="goals">' + items.map(g => {
    const p = Math.max(2, Math.round(g.value / max * 100));
    return '<div class="goal"><div class="g-top"><b>' + g.name + "</b><span>" + fmt(g.value) + "</span></div>" +
      '<div class="g-bar"><div class="g-fill" style="width:' + p + '%"></div></div></div>';
  }).join("") + "</div>";
}

/* ---------- Overview ---------- */
function renderOverview() {
  const fin = DATA.financial.all, ship = DATA.shipments;
  document.getElementById("ov-kpis").innerHTML =
    kpi("Total Sales", money(fin.kpis.sales), deltaHtml(lastDelta(fin.monthly, "sales")),
        { icon: "🏠", iconCls: "blue", spark: fin.monthly.map(m => m.sales), color: BLUE }) +
    kpi("Total Profit", money(fin.kpis.profit), deltaHtml(lastDelta(fin.monthly, "profit")),
        { icon: "💰", iconCls: "green", spark: fin.monthly.map(m => m.profit), color: GREEN }) +
    kpi("Chocolate Sales", money(ship.kpis.sales), deltaHtml(lastDelta(ship.monthly, "sales")),
        { icon: "📦", iconCls: "orange", spark: ship.monthly.map(m => m.sales), color: ORANGE }) +
    kpi("Boxes Shipped", num(ship.kpis.boxes), deltaHtml(lastDelta(ship.monthly, "boxes")),
        { icon: "📊", iconCls: "pink", spark: ship.monthly.map(m => m.boxes), color: PINK }) +
    ringCard("Margin", Math.round(fin.kpis.margin * 100), pct(fin.kpis.margin));
  flushSparks();

  makeChart("ov-fin", lineCfg("ov-fin",
    fin.monthly.map(m => m.label),
    [{ label: "Sales", data: fin.monthly.map(m => m.sales), color: BLUE },
     { label: "Profit", data: fin.monthly.map(m => m.profit), color: GREEN, axis: "y1" }],
    { second: true }));

  makeChart("ov-seg", doughnutCfg(fin.segments.map(s => s.name), fin.segments.map(s => s.sales), "money",
    { value: money(fin.kpis.sales), label: "Total sales" }));

  makeChart("ov-cnt", barCfg(fin.countries.map(c => c.name), fin.countries.map(c => c.sales)));

  makeChart("ov-ship", lineCfg("ov-ship",
    ship.monthly.map(m => m.label),
    [{ label: "Sales", data: ship.monthly.map(m => m.sales), color: BLUE }],
    {}));

  makeChart("ov-reg", doughnutCfg(ship.regions.map(r => r.name), ship.regions.map(r => r.sales), "money",
    { value: money(ship.kpis.sales), label: "Total sales" }));

  const ins = DATA.insights;
  const share = (v, s) => v + "  <span style='color:#6b7f9f'>· " + pct(s) + " of total</span>";
  document.getElementById("ov-findings").innerHTML =
    "<table><tbody>" +
    "<tr><td><b>Top segment</b></td><td class='num'>" + ins.topSegment.name + " · " + money(ins.topSegment.value) + "</td></tr>" +
    "<tr><td><b>Top country</b></td><td class='num'>" + ins.topCountry.name + " · " + money(ins.topCountry.value) + "</td></tr>" +
    "<tr><td><b>Top product (PBI)</b></td><td class='num'>" + ins.topProduct.name + " · " + money(ins.topProduct.value) + "</td></tr>" +
    "<tr><td><b>Top category</b></td><td class='num'>" + ins.topCategory.name + " · " + money(ins.topCategory.value) + "</td></tr>" +
    "<tr><td><b>Top region</b></td><td class='num'>" + ins.topRegion.name + " · " + money(ins.topRegion.value) + "</td></tr>" +
    "<tr><td><b>Top team</b></td><td class='num'>" + ins.topTeam.name + " · " + money(ins.topTeam.value) + "</td></tr>" +
    "<tr><td><b>Top sales person</b></td><td class='num'>" + ins.topPerson.name + " · " + money(ins.topPerson.value) + "</td></tr>" +
    "</tbody></table>";

  const g = DATA.insights;
  const gcard = (label, value, sub) =>
    '<div class="card insight"><div class="i-label">' + label + '</div><div class="i-value">' + value + '</div><div class="i-sub">' + sub + "</div></div>";
  document.getElementById("ov-growth").innerHTML =
    gcard("Best month (Financial)", g.bestMonth.label, money(g.bestMonth.value) + " sales") +
    gcard("Softest month (Financial)", g.worstMonth.label, money(g.worstMonth.value) + " sales") +
    gcard("Year-on-year growth", g.yoy === null ? "n/a" : (g.yoy >= 0 ? "▲ " : "▼ ") + Math.abs(g.yoy * 100).toFixed(1) + "%", money(g.y13) + " (2013) → " + money(g.y14) + " (2014)") +
    gcard("Avg order value", money(g.avgOrder), money(g.avgSalePerUnit) + " per unit sold") +
    gcard("Best shipment month", g.bestShipMonth.label, money(g.bestShipMonth.value) + " sales") +
    gcard("Avg shipment value", money(g.avgShipment), money(g.avgBoxValue) + " per box · " + pct(g.topRegionShare) + " top-region share") +
    gcard("Top product margin", pct(g.topProdMargin), g.topProduct.name) +
    gcard("Top person share", pct(g.topPersonShare), g.topPerson.name + " · " + money(g.topPerson.value));
}

/* ================= LIVE FILTERING (Power BI-style slicers) ================= */
const MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"];
const FIN_ALL = DATA.financial.all, SHIP_ALL = DATA.shipments;
let fSeg = null, fCnt = null, fReg = null, fTeam = null, fCat = null;  // null = all
let fYear = [2013, 2014];

function finRows() {
  return FIN_ALL.rows.filter(r =>
    r[0] >= fYear[0] && r[0] <= fYear[1] &&
    (!fSeg || fSeg.has(r[2])) && (!fCnt || fCnt.has(r[3])));
}
function shipRows() {
  return SHIP_ALL.rows.filter(r =>
    (!fReg || fReg.has(r[1])) && (!fTeam || fTeam.has(r[2])) && (!fCat || fCat.has(r[3])));
}

/* ---- in-browser aggregation (mirrors the Python build step) ---- */
function finAgg(rows) {
  const mo = new Map(), seg = new Map(), cnt = new Map(), prod = new Map(), band = new Map();
  const push = (map, k, s, p, u, c) => {
    let e = map.get(k); if (!e) { e = { s: 0, p: 0, u: 0, c: 0 }; map.set(k, e); }
    e.s += s; e.p += p; e.u += u; e.c += c;
  };
  for (const r of rows) {
    const [y, m, sn, cn, pn, bn, u, s, p] = r;
    push(mo, y * 100 + m, s, p, u, 1); push(seg, sn, s, p, u, 0);
    push(cnt, cn, s, p, u, 0); push(prod, pn, s, p, u, 0); push(band, bn, s, p, 0, 1);
  }
  const toList = m => [...m.entries()].map(([k, v]) => ({
    name: k, sales: +v.s.toFixed(2), profit: +v.p.toFixed(2), units: +v.u.toFixed(1),
    margin: v.s ? +(v.p / v.s).toFixed(4) : 0 })).sort((a, b) => b.sales - a.sales);
  const monthly = [...mo.entries()].sort((a, b) => a[0] - b[0]).map(([k, v]) => {
    const y = Math.floor(k / 100), m = k % 100;
    return { label: MONTHS[m - 1] + " " + (y % 100), sales: +v.s.toFixed(2), profit: +v.p.toFixed(2), units: +v.u.toFixed(1) };
  });
  const bands = [...band.entries()].map(([k, v]) => ({
    name: k, count: v.c, margin: v.s ? +(v.p / v.s).toFixed(4) : 0 })).sort((a, b) => b.margin - a.margin);
  let ts = 0, tp = 0, tu = 0;
  for (const r of rows) { ts += r[7]; tp += r[8]; tu += r[6]; }
  return {
    kpis: { sales: +ts.toFixed(2), profit: +tp.toFixed(2), units: +tu.toFixed(1), margin: ts ? tp / ts : 0, rows: rows.length },
    monthly, segments: toList(seg), countries: toList(cnt), products: toList(prod), bands,
  };
}

function shipAgg(rows) {
  const mo = new Map(), reg = new Map(), team = new Map(), cat = new Map(), prod = new Map(), geo = new Map(), person = new Map();
  const push = (map, k, s, b, p) => {
    let e = map.get(k); if (!e) { e = { s: 0, b: 0, p: 0 }; map.set(k, e); }
    e.s += s; e.b += b; e.p += p;
  };
  for (const r of rows) {
    const [ym, region, teamN, catN, geoN, prodN, personN, s, b, ep] = r;
    push(mo, ym, s, b, ep);
    if (region) push(reg, region, s, b, ep);
    push(team, teamN, s, b, ep);
    if (catN) push(cat, catN, s, b, ep);
    if (geoN) push(geo, geoN, s, b, ep);
    push(prod, prodN, s, b, ep);
    push(person, personN, s, b, ep);
  }
  const toList = (m, keys) => [...m.entries()].map(([k, v]) => {
    const it = { name: k };
    if (!keys || keys.includes("s")) it.sales = +v.s.toFixed(2);
    if (!keys || keys.includes("b")) it.boxes = +v.b.toFixed(1);
    if (!keys || keys.includes("p")) it.profit = +v.p.toFixed(2);
    return it;
  }).sort((a, b) => b.sales - a.sales);
  const monthly = [...mo.entries()].sort((a, b) => a[0] - b[0]).map(([k, v]) => {
    const y = Math.floor(k / 100), m = k % 100;
    return { label: MONTHS[m - 1] + " " + (y % 100), sales: +v.s.toFixed(2), boxes: +v.b.toFixed(1) };
  });
  let ts = 0, tb = 0, tep = 0;
  for (const r of rows) { ts += r[7]; tb += r[8]; tep += r[9]; }
  const topPeople = toList(person).slice(0, 10);
  const personTeam = new Map();
  for (const r of rows) personTeam.set(r[6], r[2]);
  topPeople.forEach(p => p.team = personTeam.get(p.name) || "Unknown");
  return {
    kpis: { sales: +ts.toFixed(2), boxes: +tb.toFixed(1), shipments: rows.length,
            products: prod.size, people: person.size, geos: geo.size,
            estProfit: +tep.toFixed(2), estMargin: ts ? tep / ts : 0 },
    monthly, regions: toList(reg), teams: toList(team), categories: toList(cat),
    geos: toList(geo).slice(0, 8), products: toList(prod).slice(0, 10), topPeople,
  };
}

/* ---- live insight bullets ---- */
function finInsightRows(rows, agg) {
  const k = agg.kpis, mo = agg.monthly;
  const best = mo.reduce((a, b) => (b.sales > a.sales ? b : a), mo[0]);
  const worst = mo.reduce((a, b) => (b.sales < a.sales ? b : a), mo[0]);
  let y13 = 0, y14 = 0;
  for (const r of rows) { if (r[0] === 2013) y13 += r[7]; else if (r[0] === 2014) y14 += r[7]; }
  const yoy = y13 ? (y14 - y13) / y13 : null;
  const topSeg = agg.segments[0], topCnt = agg.countries[0];
  const out = [
    { em: "📈", b: "Best month", t: best ? best.label + " · " + money(best.sales) : "—" },
    { em: "📉", b: "Softest month", t: worst ? worst.label + " · " + money(worst.sales) : "—" },
    { em: "🚀", b: "Year-on-year growth (2014 vs 2013)", t: yoy === null ? "n/a" : (yoy >= 0 ? "▲ " : "▼ ") + Math.abs(yoy * 100).toFixed(1) + "%" + (yoy >= 0 ? " growth" : " decline") + " — " + money(y13) + " → " + money(y14) },
    { em: "🧾", b: "Average sale per unit", t: money(k.sales / (k.units || 1)) + " across " + num(k.units) + " units" },
  ];
  if (topSeg) out.push({ em: "🏆", b: "Top segment", t: topSeg.name + " · " + pct(topSeg.sales / (k.sales || 1)) + " of filtered sales" });
  if (topCnt) out.push({ em: "🌍", b: "Top country", t: topCnt.name + " · " + money(topCnt.sales) + " · margin " + pct(topCnt.margin) });
  return out;
}
function shipInsightRows(rows, agg) {
  const k = agg.kpis, mo = agg.monthly;
  const best = mo.reduce((a, b) => (b.sales > a.sales ? b : a), mo[0]);
  const topReg = agg.regions[0], topTeam = agg.teams[0], topP = agg.topPeople[0];
  const out = [
    { em: "📈", b: "Best month", t: best ? best.label + " · " + money(best.sales) : "—" },
    { em: "🚚", b: "Average shipment value", t: money(k.sales / (k.shipments || 1)) + " across " + F.format(k.shipments) + " shipments" },
    { em: "🍫", b: "Estimated margin", t: pct(k.estMargin) + " · est. profit " + money(k.estProfit) },
  ];
  if (topReg) out.push({ em: "🗺️", b: "Top region", t: topReg.name + " · " + pct(topReg.sales / (k.sales || 1)) + " of sales" });
  if (topTeam) out.push({ em: "👥", b: "Top team", t: topTeam.name + " · " + money(topTeam.sales) });
  if (topP) out.push({ em: "⭐", b: "Top sales person", t: topP.name + " · " + money(topP.sales) + " (" + pct(topP.sales / (k.sales || 1)) + ")" });
  return out;
}
function insightList(rows) {
  return '<div class="insight-list">' + rows.map(r =>
    '<div class="insight-row"><div class="em">' + r.em + '</div><div class="it"><b>' + r.b + '</b><span>' + r.t + '</span></div></div>').join("") + "</div>";
}

/* ---------- Financial (live slicers) ---------- */
function renderFinancial() {
  const rows = finRows();
  const d = finAgg(rows);
  document.getElementById("fin-kpis").innerHTML =
    kpi("Total Sales", money(d.kpis.sales), deltaHtml(lastDelta(d.monthly, "sales")),
        { icon: "🏠", iconCls: "blue", spark: d.monthly.map(m => m.sales), color: BLUE }) +
    kpi("Total Profit", money(d.kpis.profit), deltaHtml(lastDelta(d.monthly, "profit")),
        { icon: "💰", iconCls: "green", spark: d.monthly.map(m => m.profit), color: GREEN }) +
    kpi("Units Sold", num(d.kpis.units),
        { text: d.kpis.rows + " rows · " + fYear[0] + "–" + fYear[1] },
        { icon: "📦", iconCls: "orange", spark: d.monthly.map(m => m.units), color: ORANGE }) +
    ringCard("Margin", Math.round(d.kpis.margin * 100), pct(d.kpis.margin));
  flushSparks();

  makeChart("fin-month", lineCfg("fin-month",
    d.monthly.map(m => m.label),
    [{ label: "Sales", data: d.monthly.map(m => m.sales), color: BLUE },
     { label: "Profit", data: d.monthly.map(m => m.profit), color: GREEN, axis: "y1" }],
    { second: true }));

  makeChart("fin-seg", doughnutCfg(d.segments.map(s => s.name), d.segments.map(s => s.sales), "money",
    { value: money(d.kpis.sales), label: "Total sales" }));

  makeChart("fin-cnt", barCfg(d.countries.map(c => c.name), d.countries.map(c => c.sales)));

  makeChart("fin-radar", radarCfg(
    d.countries.map(c => c.name),
    [
      { label: "Sales", data: shareNorm(d.countries, "sales") },
      { label: "Profit", data: shareNorm(d.countries, "profit") },
      { label: "Units", data: shareNorm(d.countries, "units") },
    ]));

  makeChart("fin-prod", barCfg(d.products.map(p => p.name), d.products.map(p => p.sales), { horizontal: true }));

  makeChart("fin-band", barCfg(
    d.bands.map(b => b.name + " (" + b.count + ")"),
    d.bands.map(b => b.margin),
    { fmt: "pct", colors: d.bands.map((_, i) => PALETTE[i % PALETTE.length]) }));

  document.getElementById("fin-goals").innerHTML =
    goalsHtml(d.products.slice(0, 5).map(p => ({ name: p.name, value: p.sales })), money);
  document.getElementById("fin-insights").innerHTML = insightList(finInsightRows(rows, d));
  document.getElementById("finFilterNote").textContent =
    "Showing " + F.format(rows.length) + " of " + F.format(FIN_ALL.rows.length) + " financial records";
}

/* ---------- Shipments (live slicers) ---------- */
function renderShipments() {
  const rows = shipRows();
  const d = shipAgg(rows);
  document.getElementById("ship-kpis").innerHTML =
    kpi("Total Sales", money(d.kpis.sales), deltaHtml(lastDelta(d.monthly, "sales")),
        { icon: "📦", iconCls: "blue", spark: d.monthly.map(m => m.sales), color: BLUE }) +
    kpi("Boxes Shipped", num(d.kpis.boxes), deltaHtml(lastDelta(d.monthly, "boxes")),
        { icon: "📊", iconCls: "green", spark: d.monthly.map(m => m.boxes), color: GREEN }) +
    kpi("Products Sold", d.kpis.products, { text: d.kpis.geos + " geographies · " + F.format(d.kpis.shipments) + " shipments" },
        { icon: "🏷️", iconCls: "orange", spark: d.categories.map(c => c.sales), color: ORANGE }) +
    kpi("Sales People", d.kpis.people, { text: d.teams.length + " teams" },
        { icon: "👥", iconCls: "pink", spark: d.teams.map(t => t.sales), color: PINK });
  flushSparks();

  makeChart("ship-month", lineCfg("ship-month",
    d.monthly.map(m => m.label),
    [{ label: "Sales", data: d.monthly.map(m => m.sales), color: BLUE },
     { label: "Boxes", data: d.monthly.map(m => m.boxes), color: ORANGE, axis: "y1" }],
    { second: true, fmt2: "int" }));

  makeChart("ship-reg", doughnutCfg(d.regions.map(r => r.name), d.regions.map(r => r.sales), "money",
    { value: money(d.kpis.sales), label: "Total sales" }));

  makeChart("ship-geo", barCfg(d.geos.map(g => g.name), d.geos.map(g => g.sales), { horizontal: true }));

  makeChart("ship-cat", barCfg(d.categories.map(c => c.name), d.categories.map(c => c.sales), { horizontal: true }));

  makeChart("ship-team", barCfg(d.teams.map(t => t.name), d.teams.map(t => t.sales), { horizontal: true, colors: d.teams.map((_, i) => PALETTE[i % PALETTE.length]) }));

  document.getElementById("ship-top-people").innerHTML =
    "<table><thead><tr><th>#</th><th>Sales Person</th><th>Team</th><th class='num'>Sales</th><th class='num'>Boxes</th></tr></thead><tbody>" +
    d.topPeople.map((p, i) => "<tr class='" + (i === 0 ? "first" : "") + "'><td><span class='rank'>" + (i + 1) + "</span></td><td><b>" + p.name + "</b></td><td>" + p.team + "</td><td class='num'>" + money(p.sales) + "</td><td class='num'>" + num(p.boxes) + "</td></tr>").join("") +
    "</tbody></table>";
  document.getElementById("ship-insights").innerHTML = insightList(shipInsightRows(rows, d));
  document.getElementById("shipFilterNote").textContent =
    "Showing " + F.format(rows.length) + " of " + F.format(SHIP_ALL.rows.length) + " shipment records";
}

/* ---------- Combined Report ---------- */
let combMetric = "sales";
function combRow(label, a, b) {
  return "<tr><td><b>" + label + "</b></td><td>" + a + "</td><td>" + b + "</td></tr>";
}
function renderCombined() {
  const c = DATA.combined, fin = DATA.financial.all, ship = DATA.shipments;
  const t = c.totals;
  document.getElementById("comb-kpis").innerHTML =
    kpi("Combined Revenue", money(t.sales), deltaHtml(lastDelta(fin.monthly, "sales")),
        { icon: "🏠", iconCls: "blue", spark: fin.monthly.map(m => m.sales), color: BLUE }) +
    kpi("Combined Est. Profit", money(t.profit), deltaHtml(lastDelta(fin.monthly, "profit")),
        { icon: "💰", iconCls: "green", spark: fin.monthly.map(m => m.profit), color: GREEN }) +
    kpi("Chocolate Est. Profit", money(t.chocProfit), deltaHtml(lastDelta(ship.monthly, "sales")),
        { icon: "🍫", iconCls: "orange", spark: ship.monthly.map(m => m.sales), color: ORANGE }) +
    kpi("Combined Volume", num(t.volume), deltaHtml(lastDelta(ship.monthly, "boxes")),
        { icon: "📦", iconCls: "pink", spark: ship.monthly.map(m => m.boxes), color: PINK }) +
    ringCard("Margin", Math.round(t.margin * 100), pct(t.margin));
  flushSparks();

  makeChart("comb-trend-fin", lineCfg("comb-trend-fin",
    fin.monthly.map(m => m.label),
    [{ label: "Sales", data: fin.monthly.map(m => m.sales), color: BLUE }],
    {}));

  makeChart("comb-trend-choc", lineCfg("comb-trend-choc",
    ship.monthly.map(m => m.label),
    [{ label: "Sales", data: ship.monthly.map(m => m.sales), color: ORANGE }],
    {}));

  makeChart("comb-seg", barCfg(fin.segments.map(s => s.name), fin.segments.map(s => s.sales), { horizontal: true }));

  makeChart("comb-reg", doughnutCfg(ship.regions.map(r => r.name), ship.regions.map(r => r.sales), "money",
    { value: money(ship.kpis.sales), label: "Total sales" }));

  const sorted = [...c.products].sort((a, b) => b[combMetric] - a[combMetric]);
  makeChart("comb-prod", barCfg(
    sorted.map(p => p.name),
    sorted.map(p => p[combMetric]),
    { horizontal: true, fmt: combMetric === "volume" ? "int" : "money",
      colors: sorted.map(p => p.dataset === "Power BI" ? TEAL : GOLD) }));

  makeChart("comb-margin", barCfg(fin.segments.map(s => s.name), fin.segments.map(s => s.margin), { fmt: "pct" }));

  makeChart("comb-cat", barCfg(ship.categories.map(x => x.name), ship.categories.map(x => x.profit), { horizontal: true }));

  const topFin = fin.products[0], topChoc = ship.products[0];
  document.getElementById("comb-table").innerHTML =
    "<table><thead><tr><th>Metric</th><th>Power BI Sample</th><th>Chocolate Co.</th></tr></thead><tbody>" +
    combRow("Rows analysed", F.format(fin.kpis.rows), F.format(ship.kpis.shipments)) +
    combRow("Period", "2013 – 2014", "Feb 2023 – Feb 2024") +
    combRow("Total sales", money(fin.kpis.sales), money(ship.kpis.sales)) +
    combRow("Profit / est. profit", money(fin.kpis.profit), money(ship.kpis.estProfit)) +
    combRow("Margin", pct(fin.kpis.margin), pct(ship.kpis.estMargin)) +
    combRow("Top product", topFin.name + " · " + money(topFin.sales), topChoc.name + " · " + money(topChoc.sales)) +
    combRow("Top contributor", "Segment: " + fin.segments[0].name, "Region: " + ship.regions[0].name) +
    combRow("Scale", fin.products.length + " products · 5 countries", ship.kpis.people + " people · 4 teams") +
    "</tbody></table>";
}

/* ---------- Data Explorer ---------- */
let dtFinQ = "", dtShipQ = "";
function renderExplorer() {
  const fr = finRows(), sr = shipRows();
  document.getElementById("data-kpis").innerHTML =
    kpi("Financial Records", F.format(fr.length), { text: "follows Financial slicers" }, { icon: "📊", iconCls: "blue" }) +
    kpi("Shipment Records", F.format(sr.length), { text: "follows Shipments slicers" }, { icon: "🚚", iconCls: "orange" }) +
    kpi("Financial Sales", money(finAgg(fr).kpis.sales), null, { icon: "💰", iconCls: "green" }) +
    kpi("Shipment Sales", money(shipAgg(sr).kpis.sales), null, { icon: "📦", iconCls: "pink" });
  renderExplorerTables();
}
function renderExplorerTables() {
  const fin = finRows(), ship = shipRows();
  const fq = dtFinQ.toLowerCase(), sq = dtShipQ.toLowerCase();
  const fRows = fq ? fin.filter(r => r.join(" ").toLowerCase().includes(fq)) : fin;
  const sRows = sq ? ship.filter(r => r.join(" ").toLowerCase().includes(sq)) : ship;
  document.getElementById("dtFinCount").textContent = F.format(fRows.length) + " of " + F.format(fin.length);
  document.getElementById("dtShipCount").textContent = F.format(sRows.length) + " of " + F.format(ship.length);
  const MAX = 250;
  const trunc = (n, cols) => n > MAX
    ? '<tr><td colspan="' + cols + '" style="color:var(--faint)">Showing first ' + MAX + " of " + F.format(n) + " rows — narrow your search</td></tr>" : "";
  document.getElementById("dtFinBody").innerHTML =
    trunc(fRows.length, 9) +
    fRows.slice(0, MAX).map(r => "<tr><td>" + r[0] + "</td><td>" + MONTHS[r[1] - 1] + "</td><td>" + r[2] + "</td><td>" + r[3] + "</td><td>" + r[4] + "</td><td>" + r[5] + "</td><td class='num'>" + num(r[6]) + "</td><td class='num'>" + money(r[7]) + "</td><td class='num'>" + money(r[8]) + "</td></tr>").join("");
  document.getElementById("dtShipBody").innerHTML =
    trunc(sRows.length, 10) +
    sRows.slice(0, MAX).map(r => {
      const y = Math.floor(r[0] / 100), m = r[0] % 100;
      return "<tr><td>" + MONTHS[m - 1] + " " + y + "</td><td>" + r[1] + "</td><td>" + r[2] + "</td><td>" + r[3] + "</td><td>" + r[4] + "</td><td>" + r[5] + "</td><td>" + r[6] + "</td><td class='num'>" + money(r[7]) + "</td><td class='num'>" + num(r[8]) + "</td><td class='num'>" + money(r[9]) + "</td></tr>";
    }).join("");
}

/* ---------- slicer wiring (Power BI-style chips) ---------- */
function buildChips(id, values, state, rows, idx, onChange) {
  const el = document.getElementById(id);
  const counts = new Map();
  for (const r of rows) counts.set(r[idx], (counts.get(r[idx]) || 0) + 1);
  el.innerHTML = values.map(v =>
    '<div class="chip on" data-k="' + v + '">' + v + ' <span class="n">' + F.format(counts.get(v) || 0) + '</span></div>').join("");
  el.querySelectorAll(".chip").forEach(ch => ch.addEventListener("click", () => {
    const v = ch.dataset.k;
    if (state.has(v) && state.size > 1) state.delete(v); else state.add(v);
    ch.classList.toggle("on", state.has(v));
    onChange();
  }));
}
function refreshChips(id, state) {
  document.querySelectorAll("#" + id + " .chip").forEach(ch => ch.classList.toggle("on", state.has(ch.dataset.k)));
}
function setupSlicers() {
  fSeg = new Set(FIN_ALL.segments.map(s => s.name));
  fCnt = new Set(FIN_ALL.countries.map(c => c.name));
  fReg = new Set(SHIP_ALL.regions.map(r => r.name));
  fTeam = new Set(SHIP_ALL.teams.map(t => t.name));
  fCat = new Set(SHIP_ALL.categories.map(c => c.name));
  buildChips("chip-seg", [...fSeg], fSeg, FIN_ALL.rows, 2, renderFinancial);
  buildChips("chip-cnt", [...fCnt], fCnt, FIN_ALL.rows, 3, renderFinancial);
  buildChips("chip-reg", [...fReg], fReg, SHIP_ALL.rows, 1, renderShipments);
  buildChips("chip-team", [...fTeam], fTeam, SHIP_ALL.rows, 2, renderShipments);
  buildChips("chip-cat", [...fCat], fCat, SHIP_ALL.rows, 3, renderShipments);

  const yrMin = document.getElementById("finYrMin"), yrMax = document.getElementById("finYrMax");
  const clamp = () => {
    let a = +yrMin.value || 2013, b = +yrMax.value || 2014;
    a = Math.max(2013, Math.min(2014, a)); b = Math.max(2013, Math.min(2014, b));
    if (a > b) { const t = a; a = b; b = t; }
    fYear = [a, b]; yrMin.value = a; yrMax.value = b;
  };
  yrMin.addEventListener("change", () => { clamp(); renderFinancial(); });
  yrMax.addEventListener("change", () => { clamp(); renderFinancial(); });
  document.getElementById("btnResetFin").addEventListener("click", () => {
    fSeg = new Set(FIN_ALL.segments.map(s => s.name));
    fCnt = new Set(FIN_ALL.countries.map(c => c.name));
    fYear = [2013, 2014]; yrMin.value = 2013; yrMax.value = 2014;
    refreshChips("chip-seg", fSeg); refreshChips("chip-cnt", fCnt);
    renderFinancial();
  });
  document.getElementById("btnResetShip").addEventListener("click", () => {
    fReg = new Set(SHIP_ALL.regions.map(r => r.name));
    fTeam = new Set(SHIP_ALL.teams.map(t => t.name));
    fCat = new Set(SHIP_ALL.categories.map(c => c.name));
    refreshChips("chip-reg", fReg); refreshChips("chip-team", fTeam); refreshChips("chip-cat", fCat);
    renderShipments();
  });
  document.getElementById("dtFinSearch").addEventListener("input", e => { dtFinQ = e.target.value; renderExplorerTables(); });
  document.getElementById("dtShipSearch").addEventListener("input", e => { dtShipQ = e.target.value; renderExplorerTables(); });
}

/* ---------- tab switching ---------- */
function showTab(name) {
  document.querySelectorAll(".tab").forEach(s => s.classList.toggle("active", s.id === "tab-" + name));
  document.querySelectorAll(".tab-btn").forEach(b => b.classList.toggle("active", b.dataset.tab === name));
  if (name === "overview") renderOverview();
  if (name === "financial") renderFinancial();
  if (name === "shipments") renderShipments();
  if (name === "combined") renderCombined();
  if (name === "data") renderExplorer();
}
document.querySelectorAll(".tab-btn").forEach(b => b.addEventListener("click", () => showTab(b.dataset.tab)));

document.querySelectorAll("#comb-metric-seg .seg-btn").forEach(b => b.addEventListener("click", () => {
  combMetric = b.dataset.metric;
  document.querySelectorAll("#comb-metric-seg .seg-btn").forEach(x => x.classList.toggle("active", x === b));
  renderCombined();
}));

setupSlicers();
const initialTab = new URLSearchParams(location.search).get("tab");
showTab(["overview", "financial", "shipments", "combined", "data"].includes(initialTab) ? initialTab : "overview");
</script>
</body>
</html>
"""


# --------------------------------------------------------------------------
# 4. Assemble everything
# --------------------------------------------------------------------------
def main():
    fin_all = agg_financial(load_financial())
    fin_2013 = agg_financial([r for r in load_financial() if r["year"] == 2013])
    fin_2014 = agg_financial([r for r in load_financial() if r["year"] == 2014])
    ship = agg_shipments(load_shipments())
    insights = build_insights(fin_all, ship)
    combined = build_combined(fin_all, ship)

    data = {
        "financial": {"all": fin_all, "2013": fin_2013, "2014": fin_2014},
        "shipments": ship,
        "insights": insights,
        "combined": combined,
    }

    with open(CHARTJS_FILE, "r", encoding="utf-8") as f:
        chartjs = f.read()
    # Never let the library accidentally terminate the inline <script> block.
    chartjs = chartjs.replace("</script>", "<\\/script>")

    html = TEMPLATE.replace("/*__CHARTJS__*/", chartjs)
    html = html.replace("/*__DATA__*/", json.dumps(data))

    with open(OUT, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"dashboard.html written ({os.path.getsize(OUT) / 1024:.0f} KB)")
    print("Financial KPIs :", fin_all["kpis"])
    print("Shipment KPIs  :", ship["kpis"])
    print("Combined totals:", combined["totals"])
    print("Top segment    :", insights["topSegment"])
    print("Top person     :", insights["topPerson"])
    print("Teams          :", [t["name"] for t in ship["teams"]])


if __name__ == "__main__":
    main()
